# app_hr/admin.py
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.urls import reverse, path
from django.template.response import TemplateResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import jdatetime

from .models import JobOpportunity, JobApplication


# ---------------------------------------
#   GLOBAL DELETE BUTTON
# ---------------------------------------
def admin_delete_button(obj):
    url = reverse(
        f"admin:{obj._meta.app_label}_{obj._meta.model_name}_delete",
        args=[obj.pk]
    )
    return format_html(
        "<a href='{}' style='background:#d7263d;color:white;"
        "padding:6px 10px;border-radius:8px;text-decoration:none;"
        "font-weight:bold;'>🗑</a>",
        url
    )

admin_delete_button.short_description = "حذف"


# ===========================================================
#   JOB OPPORTUNITY ADMIN
# ===========================================================
@admin.register(JobOpportunity)
class JobOpportunityAdmin(admin.ModelAdmin):

    list_display = (
        "poster_icon",
        "title",
        "position",
        "gender",
        "recruitment_status_badge",
        "is_active",
        "start_date_fa",
        "delete_button",
    )

    list_filter = (
        "is_active",
        "recruitment_status",
        "gender",
        "education_level",
    )

    search_fields = ("title", "position", "activity")
    ordering = ("ordering",)

    readonly_fields = (
        "slug",
        "start_date_en",
        "created_at_fa_display",
        "created_at_en",
        "updated_at_en",
        "poster_preview",
    )

    fieldsets = (
        ("📌 وضعیت", {"fields": ("is_active", "recruitment_status")}),
        ("🧾 اطلاعات شغل", {"fields": ("title", "position", "activity")}),
        ("🖼️ پوستر فراخوان", {"fields": ("poster", "poster_preview")}),
        ("💼 شرایط احراز", {
            "fields": (
                ("gender", "education_level"),
                ("min_age", "max_age"),
                "min_experience_years",
                "military_status_required",
            )
        }),
        ("📝 شرح موقعیت شغلی", {"fields": ("description_1", "description_2")}),
        ("🗓️ تاریخ‌ها", {"fields": ("start_date_fa", "start_date_en")}),
        ("ℹ️ اطلاعات سیستمی", {
            "fields": (
                "created_at_fa_display",
                "created_at_en",
                "updated_at_en",
                "slug",
            ),
            "classes": ("collapse",),
        }),
    )

    def created_at_fa_display(self, obj):
        if not obj.created_at_en:
            return "—"
        j_date = jdatetime.datetime.fromgregorian(datetime=obj.created_at_en)
        return j_date.strftime("%Y/%m/%d %H:%M")

    created_at_fa_display.short_description = "تاریخ ایجاد (شمسی)"

    def poster_icon(self, obj):
        if obj.poster:
            return format_html(
                "<img src='{}' width='45' height='45' style='border-radius:6px;object-fit:cover;'>",
                obj.poster.url
            )
        return "—"

    poster_icon.short_description = "پوستر"

    def poster_preview(self, obj):
        if obj.poster:
            return format_html(
                "<img src='{}' width='260' style='border-radius:10px;'>",
                obj.poster.url
            )
        return "—"

    def recruitment_status_badge(self, obj):
        colors = {"open": "#34a853", "closed": "#d93025"}
        return format_html(
            "<span style='background:{};color:white;padding:4px 12px;"
            "border-radius:14px;font-size:12px;'>"
            "{}"
            "</span>",
            colors.get(obj.recruitment_status, "#999"),
            obj.get_recruitment_status_display(),
        )

    recruitment_status_badge.short_description = "وضعیت جذب"

    def delete_button(self, obj):
        return admin_delete_button(obj)


# ===========================================================
#   JOB APPLICATION ADMIN (READ ONLY + EXCEL EXPORT)
# ===========================================================
@admin.register(JobApplication)
class JobApplicationAdmin(admin.ModelAdmin):

    actions = ["export_selected_to_excel"]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    list_display = (
        "full_name",
        "national_code",
        "mobile",
        "opportunity",
        "print_application",
        "print_resume",
        "created_at_fa_display",
    )

    list_filter = (
        "gender",
        "marital_status",
        "military_status",
        "opportunity",
    )

    search_fields = (
        "first_name",
        "last_name",
        "national_code",
        "email",
        "mobile",
    )

    # ----------------------------
    # EXCEL EXPORT ACTION
    # ----------------------------
    def export_selected_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.sheet_view.rightToLeft = True

        headers = [
            "نام",
            "نام خانوادگی",
            "نام کامل",
            "کد ملی",
            "جنسیت",
            "سن",
            "وضعیت تأهل",
            "وضعیت نظام وظیفه",
            "موبایل",
            "موبایل ضروری",
            "ایمیل",
            "عنوان شغلی",
            "فراخوان",
            "تاریخ ثبت (شمسی)",
            "IP",
        ]

        header_font = Font(bold=True)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        queryset = queryset.select_related("opportunity")

        for row_index, obj in enumerate(queryset, start=2):
            j_date = (
                jdatetime.datetime.fromgregorian(datetime=obj.created_at_en)
                .strftime("%Y/%m/%d %H:%M")
                if obj.created_at_en else "—"
            )

            row = [
                obj.first_name,
                obj.last_name,
                obj.full_name,
                obj.national_code,
                obj.get_gender_display(),
                obj.age,
                obj.get_marital_status_display(),
                obj.get_military_status_display() if obj.military_status else "—",
                obj.mobile,
                obj.mobile_support,
                obj.email,
                obj.opportunity.position,
                obj.opportunity.title,
                obj.get_status_display(),
                j_date,
                obj.ip_address or "—",
            ]

            for col, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col, value=value)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="job_applications.xlsx"'
        wb.save(response)
        return response

    export_selected_to_excel.short_description = "⬇ خروجی اکسل (موارد انتخاب‌شده)"

    # ----------------------------
    # READ ONLY VIEW
    # ----------------------------
    def get_readonly_fields(self, request, obj=None):
        return [field.name for field in self.model._meta.fields] + [
            "job_position",
            "resume_download",
            "created_at_fa_display",
        ]

    def created_at_fa_display(self, obj):
        if not obj.created_at_en:
            return "—"
        j_date = jdatetime.datetime.fromgregorian(datetime=obj.created_at_en)
        return j_date.strftime("%Y/%m/%d %H:%M")

    created_at_fa_display.short_description = "تاریخ ثبت (شمسی)"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:pk>/print/",
                self.admin_site.admin_view(self.print_report),
                name="app_hr_jobapplication_print",
            ),
        ]
        return custom_urls + urls

    def print_report(self, request, pk):
        application = JobApplication.objects.select_related("opportunity").get(pk=pk)
        return TemplateResponse(
            request,
            "RTL/print/job_application_print.html",
            {"application": application},
        )

    def job_position(self, obj):
        return obj.opportunity.position

    job_position.short_description = "عنوان شغلی"

    def resume_download(self, obj):
        if obj.resume_file:
            return format_html(
                "<a href='{}' target='_blank' download "
                "style='font-weight:bold;color:#1a73e8;'>⬇ دانلود رزومه</a>",
                obj.resume_file.url
            )
        return "—"

    resume_download.short_description = "رزومه"

    def print_application(self, obj):
        url = reverse("admin:app_hr_jobapplication_print", args=[obj.pk])
        return format_html(
            "<a href='{}' target='_blank' "
            "style='background:#4caf50;color:white;padding:6px 12px;"
            "border-radius:8px;text-decoration:none;font-weight:bold;'>"
            "🖨 گزارش</a>",
            url
        )

    print_application.short_description = "پرینت گزارش"

    def print_resume(self, obj):
        if obj.resume_file:
            return format_html(
                "<a href='{}' target='_blank' "
                "style='background:#1a73e8;color:white;padding:6px 12px;"
                "border-radius:8px;text-decoration:none;font-weight:bold;'>"
                "📄 رزومه</a>",
                obj.resume_file.url
            )
        return "—"

    print_resume.short_description = "پرینت رزومه"
